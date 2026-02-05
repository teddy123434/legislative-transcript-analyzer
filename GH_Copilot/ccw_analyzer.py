import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from docx import Document


class CCWTranscriptAnalyzer:
    """公督盟委員會逐字稿分析程式"""
    
    def __init__(self, speaker_list=None):
        self.speaker_data = defaultdict(lambda: {"count": 0, "content": [], "total_chars": 0})
        self.speaker_list = set(speaker_list) if speaker_list else None  # None 表示抓取所有，否則只抓指定名單
    
    def set_speaker_list(self, speaker_list):
        """設定要追蹤的委員名單"""
        self.speaker_list = set(speaker_list)
    
    def load_speaker_list_from_file(self, file_path):
        """
        從 txt 檔案讀取委員名單
        檔案格式：每行一個委員名字
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                speakers = [line.strip() for line in f if line.strip()]
            self.speaker_list = set(speakers)
            return speakers
        except Exception as e:
            raise Exception(f"讀取委員名單失敗: {e}")
    
    def clear_speaker_list(self):
        """清除委員限制，抓取所有委員"""
        self.speaker_list = None
    
    def clear_data(self):
        """清空所有統計數據（用於載入新檔案時重置）"""
        self.speaker_data = defaultdict(lambda: {"count": 0, "content": [], "total_chars": 0})

    # ---- 匯出輔助 ----
    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        """將檔名轉為合法的 Excel 工作表名稱 (<=31 chars, 去除非法字元)。"""
        invalid = set('[]:*?/\\')
        cleaned = ''.join(ch for ch in name if ch not in invalid)
        cleaned = cleaned.strip()
        return cleaned[:31] if cleaned else "sheet"

    def _write_summary_sheet(self, wb: Workbook, sheet_name: str, include_detail: bool = True):
        """將當前 self.speaker_data 寫入指定工作表。

        include_detail=True 時，會為每位委員建立詳細分頁；批次多檔案匯出時可關閉以避免大量分頁。
        """
        ws = wb.create_sheet(sheet_name)

        headers = ["委員名字", "發言次數", "總字數", "平均字數"]
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for speaker in sorted(self.speaker_data.keys()):
            data = self.speaker_data[speaker]
            count = data["count"]
            total_chars = data["total_chars"]
            avg_chars = round(total_chars / count, 1) if count > 0 else 0
            ws.append([speaker, count, total_chars, avg_chars])

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12

        if include_detail:
            # 詳細頁：每位委員一頁
            for speaker in sorted(self.speaker_data.keys()):
                ws_detail = wb.create_sheet(self._unique_sheet_name(wb, f"{sheet_name}-{speaker}"))
                data = self.speaker_data[speaker]
                ws_detail.append(["序號", "發言內容"])
                detail_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                detail_header_font = Font(bold=True, color="FFFFFF")
                for cell in ws_detail[1]:
                    cell.fill = detail_header_fill
                    cell.font = detail_header_font
                for idx, speech in enumerate(data["content"], 1):
                    ws_detail.append([idx, speech])
                ws_detail.column_dimensions['A'].width = 8
                ws_detail.column_dimensions['B'].width = 80

    def _unique_sheet_name(self, wb: Workbook, base: str) -> str:
        """確保工作表名稱在活頁簿內唯一。"""
        name = self._sanitize_sheet_name(base)
        if name not in wb.sheetnames:
            return name
        i = 2
        while True:
            candidate = self._sanitize_sheet_name(f"{name}-{i}")
            if candidate not in wb.sheetnames:
                return candidate
            i += 1
    
    def parse_transcript(self, text):
        """
        解析逐字稿文本
        格式: 委員名字：發言內容
        支援多行續接：沒有「名字：」的行會自動接續到上一個發言人
        """
        # 按換行符分割
        lines = text.strip().split('\n')
        
        current_speaker = None  # 記住當前發言人
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # 匹配 "名字：內容" 的格式
            match = re.match(r'^([^：:]+)[：:](.*?)$', line)
            if match:
                # 新的發言人
                speaker = match.group(1).strip()
                content = match.group(2).strip()
                
                # 如果有指定委員名單，則只記錄名單中的委員
                if self.speaker_list and speaker not in self.speaker_list:
                    current_speaker = None  # 不追蹤非名單中的人
                    continue
                
                current_speaker = speaker
                
                if content:  # 只記錄有內容的發言
                    self.speaker_data[speaker]["count"] += 1
                    self.speaker_data[speaker]["content"].append(content)
                    self.speaker_data[speaker]["total_chars"] += len(content)
            else:
                # 沒有冒號，視為續行文字，接續到當前發言人的最後一句
                if current_speaker and line:
                    # 將續文附加到該發言人的最後一句內容
                    if self.speaker_data[current_speaker]["content"]:
                        # 附加到最後一句，用換行分隔保持可讀性
                        self.speaker_data[current_speaker]["content"][-1] += "\n" + line
                        self.speaker_data[current_speaker]["total_chars"] += len(line) + 1  # +1 for newline
    
    def read_word_document(self, file_path):
        """
        讀取 Word 檔案 (.docx)
        支援格式: 委員名字：發言內容
        """
        try:
            doc = Document(file_path)
            text_content = []
            
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_content.append(paragraph.text)
            
            return '\n'.join(text_content)
        except Exception as e:
            raise Exception(f"讀取 Word 檔案失敗: {e}")
    
    def load_from_word(self, file_path):
        """
        從 Word 檔案載入並解析逐字稿
        """
        text = self.read_word_document(file_path)
        self.parse_transcript(text)
        return True
    
    def analyze_speaker(self, speaker_name):
        """取得特定委員的發言統計"""
        if speaker_name not in self.speaker_data:
            return None
        
        data = self.speaker_data[speaker_name]
        return {
            "name": speaker_name,
            "count": data["count"],
            "total_chars": data["total_chars"],
            "speeches": data["content"]
        }
    
    def get_all_speakers(self):
        """取得所有委員"""
        return list(self.speaker_data.keys())
    
    def export_to_excel(self, output_file="ccw_analysis.xlsx"):
        """輸出統計結果到 Excel"""
        wb = Workbook()
        # 移除預設空白頁，改用自訂名稱
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        self._write_summary_sheet(wb, "發言統計", include_detail=True)
        wb.save(output_file)
        print(f"✓ 統計結果已輸出到: {output_file}")

    def export_multiple_files_to_excel(self, file_paths, output_file="ccw_multi_analysis.xlsx"):
        """將多個會議紀錄分成各自的工作表匯出到同一個 Excel。

        - 每個檔案一個統計工作表（僅摘要，預設不建立每位委員的詳細分頁，避免檔案過大）。
        - 工作表名稱來自檔名，長度超過 31 會截斷，必要時自動加序號避免重複。
        - 支援 .docx 與 .txt；其他副檔名視為純文字。
        """
        if not file_paths:
            raise ValueError("file_paths 不可為空")

        wb = Workbook()
        wb.remove(wb.active)

        for file_path in file_paths:
            path_obj = Path(file_path)
            sheet_base = path_obj.stem or "sheet"

            # 為每個檔案建立新的分析器，並繼承當前名單限制
            analyzer = CCWTranscriptAnalyzer(self.speaker_list)

            try:
                if path_obj.suffix.lower() == ".docx":
                    analyzer.load_from_word(file_path)
                else:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        analyzer.parse_transcript(f.read())
            except Exception as e:
                raise Exception(f"處理檔案 {file_path} 時失敗: {e}")

            # 若該檔沒有任何發言，略過但保留提醒
            if not analyzer.get_all_speakers():
                continue

            sheet_name = analyzer._unique_sheet_name(wb, sheet_base)
            analyzer._write_summary_sheet(wb, sheet_name, include_detail=False)

        if not wb.sheetnames:
            raise Exception("沒有任何有效的會議紀錄可匯出")

        wb.save(output_file)
        print(f"✓ 已匯出 {len(wb.sheetnames)} 個工作表到: {output_file}")
    
    def print_summary(self, speaker_name=None):
        """列印統計摘要"""
        if speaker_name:
            # 列印特定委員
            data = self.analyze_speaker(speaker_name)
            if data:
                print(f"\n【{data['name']} 的發言統計】")
                print(f"發言次數: {data['count']}")
                print(f"總字數: {data['total_chars']}")
                print(f"\n發言內容:")
                for i, speech in enumerate(data['speeches'], 1):
                    print(f"{i}. {speech}")
            else:
                print(f"找不到 {speaker_name} 的發言")
        else:
            # 列印所有委員摘要
            print("\n【所有委員發言統計】")
            print(f"{'委員名字':<15} {'發言次數':>8} {'總字數':>8}")
            print("-" * 35)
            for speaker in sorted(self.speaker_data.keys()):
                data = self.speaker_data[speaker]
                print(f"{speaker:<15} {data['count']:>8} {data['total_chars']:>8}")


def main():
    """主程式"""
    analyzer = CCWTranscriptAnalyzer()
    
    # 示範：使用 readme.md 中的文本
    sample_text = """主席：條文及修正動議已經宣讀完畢，我們現在進行討論。
陳委員培瑜：沒有。
主席：提案委員有沒有要補充說明？
王委員義川：沒有。
主席：如果委員們沒有意見，機關有沒有要補充說明？
陳次長彥伯：沒有，謝謝，希望委員多多支持。
陳委員培瑜：不用！"""
    
    # 解析文本
    analyzer.parse_transcript(sample_text)
    
    # 列印摘要
    analyzer.print_summary()
    
    # 輸出特定委員
    analyzer.print_summary("陳委員培瑜")
    
    # 輸出到 Excel
    analyzer.export_to_excel("/Users/teddy/Programming/260127_CCW/ccw_analysis.xlsx")


if __name__ == "__main__":
    main()
